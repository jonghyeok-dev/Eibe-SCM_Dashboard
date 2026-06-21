"""
엑셀 양식 생성 및 파싱 모듈
- 기능별 개별 다운로드 + 전체 일괄 다운로드
- 고정 양식 헤더 + 입력 가이드 예시 행
- 현재고 엑셀 파싱
- '마스터' → 'DB' 워딩 교체
"""

import os
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── 공통 스타일 ──────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="29AD39", end_color="29AD39", fill_type="solid")
EXAMPLE_FONT = Font(italic=True, color="999999", size=10)
EXAMPLE_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D4D4D4"),
    right=Side(style="thin", color="D4D4D4"),
    top=Side(style="thin", color="D4D4D4"),
    bottom=Side(style="thin", color="D4D4D4"),
)


def _apply_sheet_style(ws, columns, example_row=None):
    """시트 헤더 스타일링 및 컬럼 너비 자동 조정"""
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        width = max(len(col_name) * 2.2, 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    if example_row:
        for col_idx, val in enumerate(example_row, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = val
            cell.font = EXAMPLE_FONT
            cell.fill = EXAMPLE_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

    ws.freeze_panes = "A2"


# ── 양식 정의 (DB 구조 반영) ────────────────────────────────────
TEMPLATE_DEFS = {
    "product_db": {
        "sheet_name": "품목DB",
        "columns": ["품목코드", "품목명", "카툰당입수량", "환율단위(USD/EUR/AUD/NZD)", "매입가(외화)"],
        "example": ["SN-001", "슈누프로1단계 800g", 12, "EUR", 8.50],
        "filename": "품목DB_양식.xlsx",
    },
    "warehouse_db": {
        "sheet_name": "창고DB",
        "columns": ["창고명", "창고타입(ONLINE/OFFLINE/BUYOUT)", "허용유통기한일수", "이관MOQ(캔)"],
        "example": ["용인 메인창고", "OFFLINE", 180, 0],
        "filename": "창고DB_양식.xlsx",
    },
    "logistics_cost": {
        "sheet_name": "물류비DB",
        "columns": ["출발창고ID", "도착창고ID", "카툰당물류비(원)"],
        "example": [1, 2, 3500],
        "filename": "물류비DB_양식.xlsx",
    },
    "order": {
        "sheet_name": "발주",
        "columns": ["발주월(YYYY-MM)", "품목코드", "발주수량(캔)"],
        "example": ["2026-06", "SN-001", 36000],
        "filename": "발주_양식.xlsx",
    },
    "production": {
        "sheet_name": "생산",
        "columns": ["구매코드", "생산코드", "발주월(YYYY-MM)", "생산수량(캔)", "품목코드"],
        "example": ["PO-2026-001", "PRD-2026-05", "2026-05", 36000, "SN-001"],
        "filename": "생산_양식.xlsx",
    },
    "inbound": {
        "sheet_name": "입고",
        "columns": [
            "인보이스번호", "BL번호", "매핑값", "구매코드", "생산코드",
            "선적일(YYYY-MM-DD)", "한국도착일(YYYY-MM-DD)", "ETA(YYYY-MM-DD)",
            "제조일자(YYYY-MM-DD)", "유통기한(YYYY-MM-DD)", "카툰수", "캔수",
            "단가(외화)", "총단가(외화)", "결제일(YYYY-MM-DD)", "인보이스발행일(YYYY-MM-DD)",
            "결제환율", "결제금액(원화)", "품목코드",
            "상태(생산국출발/해상운송중/한국도착/통관중/입고일선정중/입고완료)"
        ],
        "example": ["INV-2026-001", "BL-2026-001", "MAP-001", "PO-2026-001", "PRD-2026-05", 
                     "2026-03-01", "2026-04-15", "2026-04-10",
                     "2026-02-15", "2028-02-15", 300, 3600,
                     8.50, 2550.0, "2026-05-15", "2026-03-10",
                     1350.5, 3442275, "SN-001", "해상운송중"],
        "filename": "입고_양식.xlsx",
    },
    "inventory_snapshot": {
        "sheet_name": "현재고스냅샷",
        "columns": ["스냅샷일자(YYYY-MM-DD)", "창고이름", "품목명", "품목코드", "유통기한(YYYY-MM-DD)", "수량(캔)", "업데이트일시(YYYY-MM-DD HH:MM:SS)"],
        "example": ["2026-06-19", "용인 메인창고", "슈누프로1단계", "SN-001", "2028-05-15", 12000, "2026-06-19 15:30:00"],
        "filename": "현재고스냅샷_양식.xlsx",
    },
}

# 지원 타입 이름과 한글 표시명
TEMPLATE_LABELS = {
    "product_db": "품목DB",
    "warehouse_db": "창고DB",
    "logistics_cost": "물류비DB",
    "order": "발주",
    "production": "생산",
    "inbound": "입고",
    "inventory_snapshot": "현재고스냅샷",
}


def generate_template(template_type: str = "all") -> BytesIO:
    """고정 엑셀 양식 템플릿 생성"""
    output = BytesIO()

    if template_type == "all":
        types_to_generate = list(TEMPLATE_DEFS.keys())
    elif template_type in TEMPLATE_DEFS:
        types_to_generate = [template_type]
    else:
        raise ValueError(f"지원하지 않는 템플릿 유형: {template_type}")

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for t_type in types_to_generate:
            tdef = TEMPLATE_DEFS[t_type]
            pd.DataFrame(columns=tdef["columns"]).to_excel(
                writer, sheet_name=tdef["sheet_name"], index=False
            )
            ws = writer.sheets[tdef["sheet_name"]]
            _apply_sheet_style(ws, tdef["columns"], tdef.get("example"))

    output.seek(0)
    return output


def get_template_filename(template_type: str) -> str:
    """템플릿 타입에 해당하는 파일명 반환"""
    if template_type == "all":
        return "SCM_데이터입력양식_전체.xlsx"
    tdef = TEMPLATE_DEFS.get(template_type)
    if tdef:
        return tdef["filename"]
    return "SCM_양식.xlsx"


def parse_excel_file(file_path: str) -> dict:
    """업로드된 엑셀 파일을 읽어 각 시트별 DataFrame을 dict로 반환"""
    try:
        dfs = pd.read_excel(file_path, sheet_name=None, engine="openpyxl")
        return dfs
    except Exception as e:
        print(f"Excel parsing error: {e}")
        raise ValueError("엑셀 파일을 파싱하는 데 실패했습니다. 양식을 확인해주세요.")


def validate_dataframe(df: pd.DataFrame, expected_columns: list) -> bool:
    """DataFrame 컬럼 검증"""
    if df.empty:
        return False
    return all(col in df.columns for col in expected_columns)


def parse_inventory_excel(file_bytes: bytes) -> list:
    """현재고 스냅샷 엑셀 파싱"""
    try:
        dfs = pd.read_excel(BytesIO(file_bytes), sheet_name=None, engine="openpyxl")
    except Exception as e:
        raise ValueError(f"엑셀 파일 파싱 실패: {e}")

    sheet_name = "현재고스냅샷"
    if sheet_name not in dfs:
        # fallback: 첫 번째 시트 사용
        sheet_name = list(dfs.keys())[0]

    df = dfs[sheet_name]

    if df.empty:
        return []

    # 예시 행 제거
    tdef = TEMPLATE_DEFS.get("inventory_snapshot")
    if tdef and len(df) > 0:
        example = tdef["example"]
        first_row = df.iloc[0]
        try:
            if all(str(first_row.iloc[i]) == str(example[i]) for i in range(min(len(first_row), len(example)))):
                df = df.iloc[1:].reset_index(drop=True)
        except (IndexError, TypeError):
            pass

    results = []
    for _, row in df.iterrows():
        try:
            item = {
                "snapshot_date": str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else None,
                "warehouse_name": str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else None,
                "product_name": str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else None,
                "product_code": str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else None,
                "expiry_date": str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else None,
                "qty_cans": int(row.iloc[5]) if pd.notna(row.iloc[5]) else 0,
                "updated_at": str(row.iloc[6]).strip() if len(row) > 6 and pd.notna(row.iloc[6]) else None,
            }
            results.append(item)
        except (ValueError, IndexError):
            continue

    return results


def generate_order_plan_export(plan_data: list) -> BytesIO:
    """발주 계획 엑셀 내보내기"""
    output = BytesIO()
    columns = ["발주월", "품목코드", "품목명", "시스템제안수량", "실무자확정수량", "차이"]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if plan_data:
            df = pd.DataFrame(plan_data)
            df["차이"] = df["user_modified_qty"] - df["system_suggested_qty"]
            df.columns = columns
        else:
            df = pd.DataFrame(columns=columns)

        df.to_excel(writer, sheet_name="발주계획", index=False)
        ws = writer.sheets["발주계획"]
        _apply_sheet_style(ws, columns)

    output.seek(0)
    return output
